from __future__ import annotations

import hashlib
import re
from collections import defaultdict
from itertools import chain
from pathlib import Path
from typing import Any, Iterable, Optional

from openpyxl import load_workbook

from .models import DocumentRole, ItemRecord, RowType, WorkbookData
from .number_parser import math_error, parse_number, safe_amount
from .text_normalizer import (
    canonical_id,
    normalize_code,
    normalize_name,
    normalize_stt,
    normalize_text,
    normalize_unit,
    strip_accents,
)

COLUMN_PATTERNS: dict[str, list[tuple[str, int]]] = {
    "item_code": [("ma hieu", 10), ("mã hiệu", 10), ("ma cong tac", 9), ("mã công tác", 9), ("ky hieu", 7), ("code", 6)],
    "item_name": [("ten hang muc", 12), ("tên hạng mục", 12), ("noi dung cong viec", 11), ("nội dung công việc", 11), ("dien giai", 10), ("diễn giải", 10), ("mo ta", 8), ("mô tả", 8), ("ten vat tu", 8), ("tên vật tư", 8)],
    "unit": [("don vi tinh", 12), ("đơn vị tính", 12), ("dvt", 10), ("đvt", 10), ("don vi", 7)],
    "quantity": [("khoi luong nha thau", 14), ("khối lượng nhà thầu", 14), ("khoi luong moi thau", 13), ("khối lượng mời thầu", 13), ("khoi luong", 10), ("khối lượng", 10), ("so luong", 8), ("số lượng", 8)],
    "unit_price": [("don gia du thau", 14), ("đơn giá dự thầu", 14), ("don gia nha thau", 14), ("đơn giá nhà thầu", 14), ("don gia tong hop", 12), ("đơn giá tổng hợp", 12), ("don gia", 10), ("đơn giá", 10), ("price", 7)],
    "amount": [("thanh tien sau thue", 15), ("thành tiền sau thuế", 15), ("thanh tien truoc thue", 14), ("thành tiền trước thuế", 14), ("thanh tien", 11), ("thành tiền", 11), ("gia tri", 8), ("giá trị", 8)],
    "material": [("ma hieu quy cach", 13), ("mã hiệu quy cách", 13), ("quy cach", 10), ("quy cách", 10), ("vat tu", 9), ("vật tư", 9), ("vat lieu", 8), ("vật liệu", 8)],
    "brand": [("thuong hieu", 12), ("thương hiệu", 12), ("nhan hieu", 10), ("hãng", 8)],
    "origin": [("xuat xu", 12), ("xuất xứ", 12), ("nuoc san xuat", 9), ("country", 7)],
    "note": [("ghi chu", 10), ("ghi chú", 10), ("remark", 7)],
}

SKIP_SHEETS = {
    "tong hop", "tổng hợp", "dieu khoan", "điều khoản", "bia", "bìa",
    "muc luc", "mục lục", "chart", "dashboard",
}

_HEADER_TERMS = (
    "stt", "dien giai", "diễn giải", "noi dung cong viec", "nội dung công việc",
    "don vi", "đơn vị", "khoi luong", "khối lượng", "kl moi thau", "kl nhà thầu",
    "don gia", "đơn giá", "thanh tien", "thành tiền", "thuong hieu", "thương hiệu",
    "xuat xu", "xuất xứ", "ma hieu", "mã hiệu", "mo ta", "mô tả",
)

_FORMULA_ERROR = re.compile(r"#(?:REF!|DIV/0!|VALUE!|NAME\?|N/A|NUM!|NULL!)", re.I)
_ROMAN = re.compile(r"^(?:[IVXLCDM]+)(?:[.\-]|$)", re.I)
_ALPHA = re.compile(r"^[A-Z](?:[.\-]|$)", re.I)
_NUM_SECTION = re.compile(r"^\d+(?:\.\d+)*(?:[.\-]|$)")


def file_sha256(path: str | Path, chunk_size: int = 1024 * 1024) -> str:
    h = hashlib.sha256()
    with open(path, "rb") as fh:
        while chunk := fh.read(chunk_size):
            h.update(chunk)
    return h.hexdigest()


def list_sheets(path: str | Path) -> list[dict[str, Any]]:
    wb = load_workbook(path, read_only=True, data_only=True, keep_links=False)
    try:
        return [{"name": ws.title, "rows": ws.max_row or 0, "cols": ws.max_column or 0} for ws in wb.worksheets]
    finally:
        wb.close()


def _as_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _row_is_empty(row: list[Any]) -> bool:
    return not any(_as_text(v) for v in row)


def _flatten_header(rows: list[list[Any]], column_count: int) -> list[str]:
    """Flatten two/three level headers and approximate horizontal merged cells."""
    inherited = [""] * column_count
    parts: list[list[str]] = [[] for _ in range(column_count)]
    for row in rows:
        current = ""
        for col in range(column_count):
            text = _as_text(row[col] if col < len(row) else "")
            if text:
                current = text
                inherited[col] = text
            elif current:
                inherited[col] = current
            inherited_text = inherited[col]
            if inherited_text:
                norm = normalize_text(inherited_text)
                if norm and all(normalize_text(existing) != norm for existing in parts[col]):
                    parts[col].append(inherited_text)
    return [" | ".join(values) for values in parts]


def _header_score(flat: list[str]) -> int:
    joined = " | ".join(normalize_text(value) for value in flat)
    score = sum(3 for term in _HEADER_TERMS if normalize_text(term) in joined)
    populated = sum(bool(normalize_text(v)) for v in flat)
    if "stt" in joined:
        score += 8
    if "dien giai" in joined or "noi dung cong viec" in joined or "danh muc" in joined:
        score += 8
    if "don gia" in joined or "thanh tien" in joined:
        score += 6
    return score + min(populated, 20)


def detect_header(buffer: list[list[Any]], max_header_depth: int = 3) -> tuple[int, int, list[str]]:
    if not buffer:
        raise ValueError("Sheet không có dữ liệu")
    max_cols = max(len(row) for row in buffer)
    best: tuple[int, int, int, list[str]] | None = None
    scan_limit = min(len(buffer), 60)
    for start in range(scan_limit):
        if sum(bool(_as_text(v)) for v in buffer[start]) < 2:
            continue
        for depth in range(1, max_header_depth + 1):
            end = start + depth
            if end > scan_limit:
                break
            flat = _flatten_header(buffer[start:end], max_cols)
            score = _header_score(flat)
            # Prefer deeper two-level headers when the extra row adds useful labels.
            candidate = (score, -depth, -start, flat)
            if best is None or candidate[:3] > best[:3]:
                best = candidate
    if best is None or best[0] < 16:
        raise ValueError("Không tự nhận diện được hàng tiêu đề")
    depth, start = -best[1], -best[2]
    return start, start + depth - 1, best[3]


def _has(text: str, *parts: str) -> bool:
    norm = normalize_name(text)
    return all(normalize_name(part) in norm for part in parts)


def map_columns(flat_headers: list[str], role: DocumentRole) -> tuple[dict[int, str], dict[int, str]]:
    """Return fixed field mapping and dynamic technical columns.

    Multiple quantity/amount columns are retained explicitly. This is essential
    for real bidding sheets where KL mời thầu and KL nhà thầu chào coexist.
    """
    fixed: dict[int, str] = {}
    candidates: dict[str, list[tuple[int, int]]] = defaultdict(list)

    for col, raw in enumerate(flat_headers):
        text = strip_accents(normalize_text(raw))
        if not text:
            continue
        material_group = "thong tin ve vat lieu chinh" in text

        if re.search(r"\bstt\b", text):
            candidates["stt"].append((100, col)); continue
        if any(term in text for term in ("dien giai", "noi dung cong viec", "danh muc quat thong gio", "ten hang muc")):
            candidates["item_name"].append((100, col)); continue
        if ("don vi" in text or re.search(r"\bdvt\b", text)) and "don gia" not in text:
            candidates["unit"].append((95, col)); continue

        if "kl nha thau" in text or "khoi luong nha thau" in text:
            candidates["bid_quantity"].append((110, col)); continue
        if "kl moi thau" in text or "khoi luong moi thau" in text:
            priority = 110 if "lan 2" in text else (90 if "bim" in text else 80)
            candidates["reference_quantity"].append((priority, col)); continue

        if material_group and ("mo ta quy cach" in text or text.endswith("quy cach")):
            candidates["material"].append((110, col)); continue
        if material_group and "ma hieu" in text:
            candidates["item_code"].append((110, col)); continue
        if material_group and "thuong hieu" in text:
            candidates["brand"].append((110, col)); continue
        if material_group and "xuat xu" in text:
            candidates["origin"].append((110, col)); continue

        if "vl chinh" in text and "don gia" in text:
            candidates["price_main"].append((110, col)); continue
        if "vl phu" in text and "don gia" in text:
            candidates["price_aux"].append((110, col)); continue
        if ("nc m" in text or "nhan cong" in text) and "don gia" in text:
            candidates["price_labor"].append((110, col)); continue
        if "cf quan ly" in text or "chi phi quan ly" in text:
            candidates["price_management"].append((110, col)); continue
        if "loi nhuan" in text:
            candidates["price_profit"].append((110, col)); continue
        if "dg tong hop" in text or "don gia tong hop" in text:
            candidates["unit_price_total"].append((120, col)); continue

        if "thanh tien" in text and ("nha thau" in text or "hsdt" in text):
            candidates["bid_amount"].append((120, col)); continue
        if "thanh tien" in text and ("klmt" in text or "moi thau" in text or "boq" in text):
            # BOQ is a group parent, but row 7 differentiates KLMT/HSDT.
            priority = 115 if "klmt" in text or "moi thau" in text else 70
            candidates["reference_amount"].append((priority, col)); continue

        if "ghi chu" in text:
            candidates["note"].append((70 if col < 10 else 60, col)); continue

        # Generic files without the Hacom multi-level header.
        if "ma hieu" in text or "ma cong tac" in text:
            candidates["item_code"].append((70, col)); continue
        if "don gia" in text and not any(x in text for x in ("vl chinh", "vl phu", "nc m", "quan ly", "loi nhuan")):
            candidates["unit_price_total"].append((70, col)); continue
        if "thanh tien" in text:
            field = "bid_amount" if role is DocumentRole.HSDT else "reference_amount"
            candidates[field].append((60, col)); continue
        if "khoi luong" in text or re.search(r"\bkl\b", text):
            field = "bid_quantity" if role is DocumentRole.HSDT else "reference_quantity"
            candidates[field].append((50, col)); continue
        if "quy cach" in text or "vat tu" in text or "vat lieu" in text:
            candidates["material"].append((50, col)); continue
        if "thuong hieu" in text or "nhan hieu" in text:
            candidates["brand"].append((60, col)); continue
        if "xuat xu" in text:
            candidates["origin"].append((60, col)); continue

    used_cols: set[int] = set()
    for field, values in candidates.items():
        score, col = max(values, key=lambda pair: (pair[0], -pair[1]))
        if col not in used_cols:
            fixed[col] = field
            used_cols.add(col)

    technical: dict[int, str] = {}
    noise = (
        "cong ty", "lien danh", "khoi luong moi thau", "thong tin ve vat lieu chinh",
        "don gia chua bao gom vat", "thanh tien boq",
    )
    for col, raw in enumerate(flat_headers):
        if col in fixed:
            continue
        text = strip_accents(normalize_text(raw))
        if not text or any(text == value or text.endswith(value) for value in noise):
            continue
        if any(term in text for term in ("cong suat", "luu luong", "cot ap", "dien ap", "nguon dien", "moi chat", "do on", "hieu suat", "toc do", "nhiet do", "ky hieu", "quy cach", "tang h", "tang1", "tang2", "tang3", "tang4", "tang5", "tang23")):
            label = raw.split("|")[-1].strip() or raw.strip()
            technical[col] = label
    return fixed, technical


def _is_numbering_row(row: list[Any]) -> bool:
    """Detect the column-number legend, never ordinary priced rows."""
    values = [_as_text(value) for value in row if _as_text(value)]
    if len(values) < 4:
        return False
    first_four = []
    for value in values[:4]:
        match = re.match(r"^(\d+)", value)
        first_four.append(int(match.group(1)) if match else None)
    if first_four != [1, 2, 3, 4]:
        return False
    labels = sum(bool(re.match(r"^\d+(?:\s*=.*)?$", value)) for value in values)
    return labels / len(values) >= 0.70


def _iter_sheet_rows(ws) -> Iterable[tuple[int, list[Any]]]:
    for row_number, row in enumerate(ws.iter_rows(values_only=True), start=1):
        yield row_number, list(row)


def _section_level(stt: str, name: str) -> int:
    value = normalize_stt(stt)
    if _ALPHA.match(value):
        return 0
    if _ROMAN.match(value):
        return 1
    if _NUM_SECTION.match(value):
        return min(2 + value.count("."), 5)
    # Unnumbered headings are nested under the current top-level section.
    return 3 if len(name) < 160 else 4


def _looks_group(
    stt: str,
    name: str,
    unit: str,
    ref_qty: Optional[float],
    bid_qty: Optional[float],
    unit_price: Optional[float],
    ref_amount: Optional[float],
    bid_amount: Optional[float],
) -> bool:
    if not name:
        return False
    if any(value is not None for value in (ref_qty, bid_qty, unit_price, ref_amount, bid_amount)) or unit:
        return False
    return len(name.strip()) < 220


def _formula_errors(path: Path, selected_sheets: Optional[list[str]]) -> dict[tuple[str, int], list[str]]:
    """Collect formula errors without changing the values used for comparison."""
    result: dict[tuple[str, int], list[str]] = defaultdict(list)
    try:
        wb = load_workbook(path, read_only=True, data_only=False, keep_links=False)
        try:
            for ws in wb.worksheets:
                if selected_sheets is not None and ws.title not in selected_sheets:
                    continue
                if selected_sheets is None and normalize_name(ws.title) in {normalize_name(x) for x in SKIP_SHEETS}:
                    continue
                for row_number, row in enumerate(ws.iter_rows(values_only=True), start=1):
                    for col, value in enumerate(row, start=1):
                        text = _as_text(value)
                        if _FORMULA_ERROR.search(text):
                            result[(ws.title, row_number)].append(f"Lỗi công thức tại cột {col}: {text}")
        finally:
            wb.close()
    except Exception:
        return result
    return result


def load_workbook_items(
    path: str | Path,
    role: DocumentRole,
    bidder: str = "",
    selected_sheets: Optional[list[str]] = None,
    max_rows: int = 1_000_000,
) -> WorkbookData:
    path = Path(path)
    if path.suffix.lower() != ".xlsx":
        raise ValueError("Hiện tại hệ thống nhận .xlsx. Hãy Save As file .xls/.xlsb thành .xlsx trước khi chạy.")

    source_id = file_sha256(path)[:16]
    bidder_name = bidder or ("HSMT" if role is DocumentRole.HSMT else path.stem)
    formula_errors = _formula_errors(path, selected_sheets)
    wb = load_workbook(path, read_only=True, data_only=True, keep_links=False)
    items: list[ItemRecord] = []
    warnings: list[str] = []
    sheet_info: list[dict[str, Any]] = []
    totals: dict[str, float] = defaultdict(float)

    try:
        for ws in wb.worksheets:
            sheet_norm = normalize_name(ws.title)
            if selected_sheets is not None and ws.title not in selected_sheets:
                continue
            if selected_sheets is None and any(normalize_name(skip) == sheet_norm for skip in SKIP_SHEETS):
                continue

            iterator = _iter_sheet_rows(ws)
            buffer: list[tuple[int, list[Any]]] = []
            for _ in range(90):
                try:
                    buffer.append(next(iterator))
                except StopIteration:
                    break
            if not buffer:
                continue

            try:
                start, end, flat_headers = detect_header([row for _, row in buffer])
                mapping, technical_columns = map_columns(flat_headers, role)
            except ValueError as exc:
                warnings.append(f"Sheet '{ws.title}': {exc}")
                continue

            if "item_name" not in mapping.values():
                warnings.append(f"Sheet '{ws.title}': không xác định được cột tên hạng mục")
                continue

            sheet_info.append({
                "sheet": ws.title,
                "header_start": buffer[start][0],
                "header_end": buffer[end][0],
                "mapped_columns": {flat_headers[col]: field for col, field in mapping.items()},
                "technical_columns": {flat_headers[col]: label for col, label in technical_columns.items()},
            })

            section_stack: list[str] = []
            section_code_stack: list[str] = []
            current_parent_detail = ""
            current_parent_code = ""
            empty_streak = 0
            processed = 0

            for row_number, row in chain(buffer, iterator):
                if row_number <= buffer[end][0]:
                    continue
                if _is_numbering_row(row):
                    continue
                if _row_is_empty(row):
                    empty_streak += 1
                    if empty_streak > 250:
                        break
                    continue
                empty_streak = 0
                processed += 1
                if processed > max_rows:
                    warnings.append(f"Sheet '{ws.title}' vượt giới hạn {max_rows:,} dòng")
                    break

                values: dict[str, Any] = {field: (row[col] if col < len(row) else None) for col, field in mapping.items()}
                stt = _as_text(values.get("stt"))
                name = _as_text(values.get("item_name"))
                code = _as_text(values.get("item_code"))
                unit = _as_text(values.get("unit"))
                if not name and not code:
                    continue

                ref_qty = parse_number(values.get("reference_quantity"))
                bid_qty = parse_number(values.get("bid_quantity"))
                price_main = parse_number(values.get("price_main"))
                price_aux = parse_number(values.get("price_aux"))
                price_labor = parse_number(values.get("price_labor"))
                price_management = parse_number(values.get("price_management"))
                price_profit = parse_number(values.get("price_profit"))
                unit_price_total = parse_number(values.get("unit_price_total"))
                if unit_price_total is None:
                    components = [price_main, price_aux, price_labor, price_management, price_profit]
                    if any(value is not None for value in components):
                        unit_price_total = sum(value or 0.0 for value in components)

                raw_ref_amount = parse_number(values.get("reference_amount"))
                raw_bid_amount = parse_number(values.get("bid_amount"))
                reference_amount = safe_amount(ref_qty, unit_price_total, raw_ref_amount)
                bid_amount = safe_amount(bid_qty, unit_price_total, raw_bid_amount)

                summary = (
                    not unit
                    and ref_qty is None
                    and bid_qty is None
                    and unit_price_total is None
                    and (raw_ref_amount is not None or raw_bid_amount is not None)
                )
                group = _looks_group(stt, name, unit, ref_qty, bid_qty, unit_price_total, raw_ref_amount, raw_bid_amount)
                normalized_stt = normalize_stt(stt)

                if summary:
                    row_type = RowType.SUMMARY
                    current_parent_detail = ""
                    current_parent_code = ""
                elif group:
                    level = _section_level(stt, name)
                    if len(section_stack) <= level:
                        section_stack.extend([""] * (level + 1 - len(section_stack)))
                        section_code_stack.extend([""] * (level + 1 - len(section_code_stack)))
                    section_stack[level] = name
                    section_code_stack[level] = normalized_stt
                    del section_stack[level + 1:]
                    del section_code_stack[level + 1:]
                    current_parent_detail = ""
                    current_parent_code = ""
                    row_type = RowType.GROUP
                else:
                    # Rows with an explicit STT are principal BOQ lines. Rows without
                    # one are treated as components under the immediately preceding
                    # principal line, preventing components from different cabinets
                    # from being matched to each other.
                    row_type = RowType.DETAIL if normalized_stt else RowType.COMPONENT
                    if row_type is RowType.DETAIL:
                        current_parent_detail = name
                        current_parent_code = normalized_stt or normalize_code(code)

                path_parts = tuple(value for value in section_stack if value)
                code_parts = tuple(value for value in section_code_stack if value)
                if row_type is RowType.COMPONENT and current_parent_detail:
                    path_parts = path_parts + (current_parent_detail,)
                    code_parts = code_parts + (current_parent_code,)

                technical_specs: dict[str, Any] = {}
                for col, label in technical_columns.items():
                    value = row[col] if col < len(row) else None
                    if _as_text(value):
                        technical_specs[label] = value

                quality: list[str] = list(formula_errors.get((ws.title, row_number), []))
                raw_values = [row[col] for col in range(min(len(row), len(flat_headers)))]
                for col, value in enumerate(raw_values, start=1):
                    if isinstance(value, str) and _FORMULA_ERROR.search(value):
                        quality.append(f"Lỗi dữ liệu tại cột {col}: {value}")

                if row_type in {RowType.DETAIL, RowType.COMPONENT}:
                    quantity = bid_qty if role is DocumentRole.HSDT else ref_qty
                    amount = bid_amount if role is DocumentRole.HSDT else reference_amount
                    # Principal BOQ rows must carry quantity, unit price and amount.
                    # Component/specification rows often intentionally omit prices;
                    # treating those blanks as errors would flood the report with
                    # false positives and hide the real bidder anomalies.
                    if row_type is RowType.DETAIL:
                        if quantity is None:
                            quality.append("Thiếu khối lượng")
                        if unit_price_total is None:
                            quality.append("Thiếu đơn giá tổng hợp")
                        if amount is None:
                            quality.append("Thiếu thành tiền")
                    elif unit and quantity is None:
                        quality.append("Thiếu khối lượng cấu thành")

                    err = math_error(quantity, unit_price_total, amount)
                    if err is not None:
                        quality.append(f"Sai phép tính KL×ĐG, lệch {err:,.0f}")
                    component_values = [price_main, price_aux, price_labor, price_management, price_profit]
                    if unit_price_total is not None and any(v is not None for v in component_values):
                        component_sum = sum(v or 0.0 for v in component_values)
                        scale = max(abs(unit_price_total), abs(component_sum), 1.0)
                        if abs(component_sum - unit_price_total) / scale > 0.005:
                            quality.append(f"Tổng 5 thành phần giá lệch ĐG tổng hợp {component_sum - unit_price_total:+,.0f}")

                normalized_path = normalize_name(" | ".join(path_parts))
                normalized_name = normalize_name(name)
                structural_key = "::".join(filter(None, [
                    normalize_name(ws.title),
                    normalized_path,
                    normalized_stt or normalize_code(code),
                    normalized_name if row_type is RowType.COMPONENT else "",
                    row_type.value,
                ]))

                item = ItemRecord(
                    source_id=source_id,
                    role=role,
                    bidder=bidder_name,
                    workbook=path.name,
                    sheet=ws.title,
                    row_number=row_number,
                    stt=stt,
                    item_code=code,
                    item_name=name,
                    unit=unit,
                    reference_quantity=ref_qty,
                    bid_quantity=bid_qty,
                    price_main=price_main,
                    price_aux=price_aux,
                    price_labor=price_labor,
                    price_management=price_management,
                    price_profit=price_profit,
                    unit_price_total=unit_price_total,
                    reference_amount=reference_amount,
                    bid_amount=bid_amount,
                    material=_as_text(values.get("material")),
                    brand=_as_text(values.get("brand")),
                    origin=_as_text(values.get("origin")),
                    note=_as_text(values.get("note")),
                    technical_specs=technical_specs,
                    section_path=path_parts,
                    section_codes=code_parts,
                    row_type=row_type,
                    raw={flat_headers[col]: _as_text(row[col] if col < len(row) else None) for col in range(len(flat_headers))},
                    normalized_stt=normalized_stt,
                    normalized_code=normalize_code(code),
                    normalized_name=normalized_name,
                    normalized_unit=normalize_unit(unit),
                    normalized_path=normalized_path,
                    structural_key=structural_key,
                    data_quality_flags=list(dict.fromkeys(quality)),
                )
                items.append(item)
                if row_type is RowType.DETAIL and item.amount is not None:
                    totals[ws.title] += item.amount

    finally:
        wb.close()

    # Duplicate codes are retained but flagged only when the descriptions differ.
    by_code: dict[tuple[str, str], list[ItemRecord]] = defaultdict(list)
    for item in items:
        if item.is_comparable and item.normalized_code:
            by_code[(normalize_name(item.sheet), item.normalized_code)].append(item)
    for (_, code), duplicated in by_code.items():
        distinct_names = {item.normalized_name for item in duplicated if item.normalized_name}
        if len(duplicated) > 1 and len(distinct_names) > 1:
            rows = ", ".join(str(item.row_number) for item in duplicated[:20])
            flag = f"Mã hiệu trùng nhưng mô tả khác nhau ({code}), các dòng: {rows}"
            for item in duplicated:
                item.data_quality_flags.append(flag)
            warnings.append(f"Sheet '{duplicated[0].sheet}': {flag}")

    # Structural duplicates are reported, never silently collapsed.
    by_structure: dict[str, list[ItemRecord]] = defaultdict(list)
    for item in items:
        if item.is_comparable and item.structural_key:
            by_structure[item.structural_key].append(item)
    for key, duplicated in by_structure.items():
        if len(duplicated) > 1:
            warnings.append(
                f"Sheet '{duplicated[0].sheet}': khóa cấu trúc lặp {len(duplicated)} lần; "
                f"giữ nguyên từng dòng ({', '.join(str(x.row_number) for x in duplicated[:10])})"
            )

    if not items:
        warnings.append("Không đọc được hạng mục dữ liệu nào từ workbook")
    return WorkbookData(
        path=path,
        role=role,
        bidder=bidder_name,
        items=items,
        warnings=warnings,
        sheet_info=sheet_info,
        totals=dict(totals),
    )

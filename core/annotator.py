from __future__ import annotations

from collections import defaultdict
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .models import ComparedItem, FieldDifference, Severity, WorkbookData

_FILL = {
    Severity.INFO: PatternFill("solid", fgColor="DDEBF7"),
    Severity.REVIEW: PatternFill("solid", fgColor="FFF2CC"),
    Severity.WARNING: PatternFill("solid", fgColor="FCE4D6"),
    Severity.CRITICAL: PatternFill("solid", fgColor="F4CCCC"),
}
_FONT = {
    Severity.INFO: "1F4E78",
    Severity.REVIEW: "7F6000",
    Severity.WARNING: "C65911",
    Severity.CRITICAL: "9C0006",
}
_RANK = {Severity.OK: 0, Severity.INFO: 1, Severity.REVIEW: 2, Severity.WARNING: 3, Severity.CRITICAL: 4}
_THIN = Side(style="thin", color="D9E1F2")


def _worst(rows: Iterable[ComparedItem]) -> Severity:
    return max((row.severity for row in rows), key=lambda s: _RANK[s], default=Severity.OK)


def _sheet_meta(workbook: WorkbookData) -> dict[str, dict]:
    return {str(info.get("sheet")): info for info in workbook.sheet_info}


def _field_key(diff: FieldDifference) -> str | None:
    field = diff.field.lower()
    if "tên hạng mục" in field or field == "hạng mục":
        return "item_name"
    if "mã hiệu" in field:
        return "item_code"
    if "đơn vị" in field:
        return "unit"
    if "khối lượng mời thầu" in field:
        return "reference_quantity"
    if "khối lượng nhà thầu" in field:
        return "bid_quantity"
    if "đơn giá tổng hợp" in field:
        return "unit_price_total"
    if "thành tiền theo klmt" in field:
        return "reference_amount"
    if "thành tiền nhà thầu" in field:
        return "bid_amount"
    if "vl chính" in field:
        return "price_main"
    if "vl phụ" in field:
        return "price_aux"
    if "nc & máy" in field or "nc&m" in field:
        return "price_labor"
    if "quản lý" in field:
        return "price_management"
    if "lợi nhuận" in field:
        return "price_profit"
    if "vật tư" in field or "quy cách" in field:
        return "material"
    if "thương hiệu" in field:
        return "brand"
    if "xuất xứ" in field:
        return "origin"
    return None


def _comment_text(row: ComparedItem, max_chars: int = 6000) -> str:
    lines = [
        f"AI đánh dấu: {row.severity.value}",
        f"Điểm bất thường: {row.anomaly_score:.1f}/100",
        f"Ghép với PL01: {row.match.kind.value}, độ tin cậy {row.match.score:.1%}",
    ]
    if row.pl2_category:
        lines.append(f"Nhóm PL02: {row.pl2_category}")
        lines.append(f"Yêu cầu PL02: {row.pl2_requirement}")
        lines.append(f"Trạng thái PL02: {row.pl2_status}")
    if row.flags:
        lines.append("Lý do:")
        lines.extend(f"- {flag}" for flag in row.flags)
    text = "\n".join(lines)
    return text[:max_chars]


def _append_comment(cell, text: str) -> None:
    if cell.comment and cell.comment.text:
        text = cell.comment.text + "\n\n--- HSMT Enterprise AI ---\n" + text
    cell.comment = Comment(text, "HSMT Enterprise AI")


def _prepare_review_sheet(wb, bidder: str):
    for name in ("AI_TONG_QUAN", "AI_KIEM_TRA"):
        if name in wb.sheetnames:
            del wb[name]
    summary = wb.create_sheet("AI_TONG_QUAN", 0)
    review = wb.create_sheet("AI_KIEM_TRA", 1)

    summary["A1"] = "KẾT QUẢ KIỂM TRA HỒ SƠ CHÀO GIÁ"
    summary["A2"] = f"Nhà thầu: {bidder}"
    summary["A4"] = "Màu"
    summary["B4"] = "Ý nghĩa"
    legend = [
        (Severity.REVIEW, "Cần chuyên viên xác nhận"),
        (Severity.WARNING, "Sai lệch đáng kể"),
        (Severity.CRITICAL, "Thiếu, sai công thức hoặc chênh lệch nghiêm trọng"),
    ]
    for index, (severity, meaning) in enumerate(legend, 5):
        summary.cell(index, 1, severity.value)
        summary.cell(index, 1).fill = _FILL[severity]
        summary.cell(index, 1).font = Font(name="Arial", bold=True, color=_FONT[severity])
        summary.cell(index, 2, meaning)
    summary["A10"] = "Lưu ý"
    summary["B10"] = (
        "Các đánh dấu là tín hiệu hỗ trợ rà soát. Thương hiệu ngoài Phụ lục 02 không tự động bị loại; "
        "cần kiểm tra tài liệu chứng minh tương đương hoặc tốt hơn. Khác tên sheet chỉ được ghi chú, "
        "không được tính là cảnh báo khi hạng mục đã khớp."
    )
    summary.column_dimensions["A"].width = 25
    summary.column_dimensions["B"].width = 95
    summary.merge_cells("A1:H1")
    summary["A1"].font = Font(name="Arial", bold=True, size=16, color="FFFFFF")
    summary["A1"].fill = PatternFill("solid", fgColor="17365D")
    summary["A1"].alignment = Alignment(horizontal="center")

    headers = [
        "Mức độ", "Sheet gốc", "Dòng gốc", "STT", "Hạng mục PL01", "Hạng mục nhà thầu",
        "Thông số", "Giá trị yêu cầu/nhóm", "Giá trị nhà thầu", "Chênh lệch", "Chênh lệch (%)",
        "Lý do", "Liên kết tới dòng gốc",
    ]
    for col, header in enumerate(headers, 1):
        cell = review.cell(1, col, header)
        cell.fill = PatternFill("solid", fgColor="1F4E78")
        cell.font = Font(name="Arial", bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(bottom=_THIN)
    widths = [18, 24, 12, 14, 48, 48, 30, 35, 35, 18, 18, 80, 22]
    for col, width in enumerate(widths, 1):
        review.column_dimensions[get_column_letter(col)].width = width
    review.freeze_panes = "A2"
    review.auto_filter.ref = f"A1:M1"
    return summary, review


def annotate_bidder_workbook(
    source_path: str | Path,
    output_path: str | Path,
    bidder_workbook: WorkbookData,
    rows: list[ComparedItem],
) -> str:
    """Create an annotated copy while preserving original formulas and layout.

    The original workbook is never overwritten. Each source sheet receives two
    additional columns (AI MỨC ĐỘ, AI LÝ DO, AI GHI CHÚ), and a front review sheet links
    directly back to the source row.
    """
    source_path, output_path = Path(source_path), Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb = load_workbook(source_path, data_only=False, keep_links=True)
    try:
        summary_ws, review_ws = _prepare_review_sheet(wb, bidder_workbook.bidder)
        meta = _sheet_meta(bidder_workbook)
        grouped_by_location: dict[tuple[str, int], list[ComparedItem]] = defaultdict(list)
        notes_by_location: dict[tuple[str, int], list[str]] = defaultdict(list)
        for compared in rows:
            if compared.candidate is None:
                continue
            location = (compared.candidate.sheet, compared.candidate.row_number)
            if compared.severity is not Severity.OK:
                grouped_by_location[location].append(compared)
            for note in compared.notes:
                if note not in notes_by_location[location]:
                    notes_by_location[location].append(note)

        review_row = 2
        counts = defaultdict(int)
        for compared in rows:
            if compared.severity is Severity.OK:
                continue
            counts[compared.severity.value] += 1
            candidate, reference = compared.candidate, compared.reference
            diffs = compared.differences
            fields = list(dict.fromkeys(diff.field for diff in diffs))
            ref_values = list(dict.fromkeys(str(diff.reference_value) for diff in diffs if diff.reference_value not in (None, "")))
            cand_values = list(dict.fromkeys(str(diff.candidate_value) for diff in diffs if diff.candidate_value not in (None, "")))
            deltas = [diff.delta for diff in diffs if isinstance(diff.delta, (int, float))]
            delta_pcts = [diff.delta_pct for diff in diffs if isinstance(diff.delta_pct, (int, float))]
            reasons = list(dict.fromkeys(compared.flags or [compared.match.reason]))
            values = [
                compared.severity.value,
                candidate.sheet if candidate else "",
                candidate.row_number if candidate else None,
                candidate.stt if candidate else (reference.stt if reference else ""),
                reference.item_name if reference else "",
                candidate.item_name if candidate else "",
                " | ".join(fields),
                " | ".join(ref_values)[:8000],
                " | ".join(cand_values)[:8000],
                max(deltas, key=abs) if deltas else None,
                max(delta_pcts, key=abs) if delta_pcts else None,
                " | ".join(reasons)[:30000],
                "Mở dòng gốc" if candidate else "Không có dòng để mở",
            ]
            for col, value in enumerate(values, 1):
                cell = review_ws.cell(review_row, col, value)
                cell.alignment = Alignment(vertical="top", wrap_text=col in {5, 6, 7, 8, 9, 12})
                if col == 1:
                    cell.fill = _FILL.get(compared.severity, PatternFill())
                    cell.font = Font(name="Arial", bold=True, color=_FONT.get(compared.severity, "000000"))
                if col == 11 and isinstance(value, (int, float)):
                    cell.number_format = "0.00%"
                if col == 10 and isinstance(value, (int, float)):
                    cell.number_format = "#,##0.000;[Red]-#,##0.000"
            if candidate:
                escaped = candidate.sheet.replace("'", "''")
                review_ws.cell(review_row, 13).hyperlink = f"#'{escaped}'!A{candidate.row_number}"
                review_ws.cell(review_row, 13).style = "Hyperlink"
            review_row += 1

        # Add exact spreadsheet formula/external-link issues, including cells
        # outside parsed BOQ rows. These are detected by direct OOXML scanning.
        for issue in bidder_workbook.formula_issues:
            kind = str(issue.get("kind", ""))
            severity = Severity.CRITICAL if kind == "FORMULA_ERROR" else Severity.REVIEW
            counts[severity.value] += 1
            sheet_name = str(issue.get("sheet", ""))
            row_number = int(issue.get("row", 0) or 0)
            cell_ref = str(issue.get("cell", ""))
            message = str(issue.get("message", ""))
            formula = str(issue.get("formula", ""))
            values = [
                severity.value, sheet_name, row_number, "", "", "",
                f"Lỗi Excel tại ô {cell_ref}", formula, str(issue.get("value", "")),
                None, None, message, "Mở ô lỗi",
            ]
            for col, value in enumerate(values, 1):
                cell = review_ws.cell(review_row, col, value)
                cell.alignment = Alignment(vertical="top", wrap_text=col in {7, 8, 9, 12})
                if col == 1:
                    cell.fill = _FILL[severity]
                    cell.font = Font(name="Arial", bold=True, color=_FONT[severity])
            if sheet_name and cell_ref and sheet_name in wb.sheetnames:
                escaped = sheet_name.replace("'", "''")
                review_ws.cell(review_row, 13).hyperlink = f"#'{escaped}'!{cell_ref}"
                review_ws.cell(review_row, 13).style = "Hyperlink"
            review_row += 1

        summary_ws["A12"] = "Số dòng cần kiểm tra"
        summary_ws["B12"] = counts.get(Severity.REVIEW.value, 0)
        summary_ws["A13"] = "Số dòng cảnh báo"
        summary_ws["B13"] = counts.get(Severity.WARNING.value, 0)
        summary_ws["A14"] = "Số dòng bất thường"
        summary_ws["B14"] = counts.get(Severity.CRITICAL.value, 0)
        summary_ws["A16"] = "Tổng dòng trong AI_KIEM_TRA"
        summary_ws["B16"] = max(0, review_row - 2)
        summary_ws["A17"] = "Lỗi công thức/liên kết Excel"
        summary_ws["B17"] = len(bidder_workbook.formula_issues)
        summary_ws["A18"] = "External links trong workbook"
        summary_ws["B18"] = bidder_workbook.external_link_count

        # Fix the AI columns once per sheet. Recomputing from ws.max_column after
        # every row would keep appending new columns and make the file explode.
        ai_columns: dict[str, tuple[int, int, int, int, dict[str, int]]] = {}
        for sheet_name, info in meta.items():
            if sheet_name not in wb.sheetnames:
                continue
            ws = wb[sheet_name]
            fields = {str(k): int(v) for k, v in (info.get("field_columns") or {}).items()}
            header_row = int(info.get("header_end") or 1)
            original_max = max(int(info.get("max_column") or 1), int(ws.max_column or 1))
            ai_sev_col, ai_reason_col, ai_note_col = original_max + 2, original_max + 3, original_max + 4
            ai_columns[sheet_name] = (ai_sev_col, ai_reason_col, ai_note_col, header_row, fields)
            ws.cell(header_row, ai_sev_col, "AI MỨC ĐỘ")
            ws.cell(header_row, ai_reason_col, "AI LÝ DO")
            ws.cell(header_row, ai_note_col, "AI GHI CHÚ")
            for col in (ai_sev_col, ai_reason_col, ai_note_col):
                ws.cell(header_row, col).fill = PatternFill("solid", fgColor="17365D")
                ws.cell(header_row, col).font = Font(name="Arial", bold=True, color="FFFFFF")
                ws.cell(header_row, col).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            ws.column_dimensions[get_column_letter(ai_sev_col)].width = 18
            ws.column_dimensions[get_column_letter(ai_reason_col)].width = 80
            ws.column_dimensions[get_column_letter(ai_note_col)].width = 70

        # Informational notes are written for both OK and anomalous rows.  They
        # never enter AI_KIEM_TRA and never change the row colour or severity.
        for (sheet_name, row_number), notes in notes_by_location.items():
            if sheet_name not in wb.sheetnames or sheet_name not in ai_columns:
                continue
            ws = wb[sheet_name]
            _, _, ai_note_col, _, _ = ai_columns[sheet_name]
            note_text = " | ".join(dict.fromkeys(notes))[:32000]
            ws.cell(row_number, ai_note_col, note_text)
            ws.cell(row_number, ai_note_col).alignment = Alignment(vertical="top", wrap_text=True)

        for (sheet_name, row_number), location_rows in grouped_by_location.items():
            if sheet_name not in wb.sheetnames or sheet_name not in ai_columns:
                continue
            ws = wb[sheet_name]
            ai_sev_col, ai_reason_col, _, _, fields = ai_columns[sheet_name]
            severity = _worst(location_rows)
            reasons: list[str] = []
            for compared in location_rows:
                reasons.extend(compared.flags)
            reasons = list(dict.fromkeys(reasons))
            reason_text = " | ".join(reasons)[:32000]

            ws.cell(row_number, ai_sev_col, severity.value)
            ws.cell(row_number, ai_reason_col, reason_text)
            for col in (ai_sev_col, ai_reason_col):
                ws.cell(row_number, col).fill = _FILL.get(severity, PatternFill())
                ws.cell(row_number, col).font = Font(name="Arial", color=_FONT.get(severity, "000000"), bold=col == ai_sev_col)
                ws.cell(row_number, col).alignment = Alignment(vertical="top", wrap_text=True)

            item_col = fields.get("item_name", 2 if ws.max_column >= 2 else 1)
            anchor = ws.cell(row_number, item_col)
            anchor.fill = _FILL.get(severity, PatternFill())
            anchor.font = Font(
                name="Arial", size=anchor.font.sz, bold=True, italic=anchor.font.italic,
                color=_FONT.get(severity, "000000"), underline=anchor.font.underline,
            )

            # Highlight the actual value cells, while keeping formula/value intact.
            for compared in location_rows:
                for diff in compared.differences:
                    key = _field_key(diff)
                    col = fields.get(key) if key else None
                    if col:
                        ws.cell(row_number, col).fill = _FILL.get(diff.severity, _FILL.get(severity, PatternFill()))

        # Highlight exact error cells and also write the AI reason on that row.
        for issue in bidder_workbook.formula_issues:
            sheet_name = str(issue.get("sheet", ""))
            cell_ref = str(issue.get("cell", ""))
            row_number = int(issue.get("row", 0) or 0)
            kind = str(issue.get("kind", ""))
            message = str(issue.get("message", ""))
            severity = Severity.CRITICAL if kind == "FORMULA_ERROR" else Severity.REVIEW
            if sheet_name not in wb.sheetnames or not cell_ref:
                continue
            ws = wb[sheet_name]
            target = ws[cell_ref]
            target.fill = _FILL[severity]
            target.font = Font(
                name="Arial", size=target.font.sz, bold=True,
                italic=target.font.italic, color=_FONT[severity], underline=target.font.underline,
            )
            _append_comment(target, message)
            if sheet_name in ai_columns and row_number > 0:
                ai_sev_col, ai_reason_col, _, _, _ = ai_columns[sheet_name]
                existing = str(ws.cell(row_number, ai_reason_col).value or "").strip()
                combined = " | ".join(value for value in (existing, message) if value)[:32000]
                current_severity = str(ws.cell(row_number, ai_sev_col).value or "")
                if severity is Severity.CRITICAL or not current_severity:
                    ws.cell(row_number, ai_sev_col, severity.value)
                ws.cell(row_number, ai_reason_col, combined)
                for col in (ai_sev_col, ai_reason_col):
                    ws.cell(row_number, col).fill = _FILL[severity]
                    ws.cell(row_number, col).font = Font(name="Arial", color=_FONT[severity], bold=col == ai_sev_col)
                    ws.cell(row_number, col).alignment = Alignment(vertical="top", wrap_text=True)

        review_ws.auto_filter.ref = f"A1:M{max(1, review_row - 1)}"
        wb.calculation.fullCalcOnLoad = True
        wb.calculation.forceFullCalc = True
        wb.save(output_path)
    finally:
        wb.close()
    return str(output_path)


"""Kiểm tra file tổng hợp độc lập 'Bảng chào giá tổng hợp' - hệ thống tự sinh
theo đúng format file tổng hợp thực tế: mỗi hạng mục một sheet, các nhà thầu
xếp cạnh nhau theo block, và đánh dấu trực tiếp lên ô 'ĐG tổng hợp' / 'Thành
tiền NT chào' khi lệch nhiều. File KHÔNG có cột phân tích (Mức độ, Điểm bất thường).
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook, load_workbook

from core.config import EnterpriseConfig
from core.pipeline import compare_bidder_files
from core.reporter import export_consolidated_summary


def _cfg() -> EnterpriseConfig:
    cfg = EnterpriseConfig()
    cfg.enable_semantic_matching = False
    cfg.enable_reranker = False
    return cfg


def _bidder_book(path: Path, unit_price: float, sheet_title: str = "1. HT điện") -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_title
    ws.append(["STT", "Mã hiệu", "Tên hạng mục", "ĐVT", "Khối lượng", "Đơn giá tổng hợp", "Thành tiền"])
    ws.append(["1", "M-01", "Tủ điện tổng LV-G.1", "Cái", 1, unit_price, unit_price])
    wb.save(path)


def _build_summary(tmp_path: Path, prices: dict[str, float]) -> Path:
    bidder_files = []
    for name, price in prices.items():
        path = tmp_path / f"{name}.xlsx"
        _bidder_book(path, price)
        bidder_files.append((name, path))
    result = compare_bidder_files(bidder_files, config=_cfg())
    summary_path = tmp_path / "Bang_tong_hop.xlsx"
    export_consolidated_summary(result, summary_path)
    return summary_path


def _leaf_columns(ws):
    """Map {tên cột con -> [chỉ số cột Excel]} đọc từ tiêu đề tầng 3 (row 4)."""
    result = {}
    for cell in ws[4]:
        if cell.value:
            result.setdefault(str(cell.value).replace("\n", " ").strip(), []).append(cell.column)
    return result


def test_summary_has_one_sheet_per_hangmuc_with_side_by_side_blocks(tmp_path: Path):
    prices = {"NT A": 100, "NT B": 105, "NT C": 95, "NT Lệch": 500}
    summary_path = _build_summary(tmp_path, prices)

    wb = load_workbook(summary_path)
    # Hạng mục gốc là "1. HT điện" -> đúng một sheet mang tên đó.
    assert "1. HT điện" in wb.sheetnames
    ws = wb["1. HT điện"]

    # Tầng 1 (row 2) phải có đủ tên cả 4 nhà thầu.
    tier1 = " ".join(str(c.value) for c in ws[2] if c.value)
    for name in prices:
        assert name in tier1

    # Không có cột phân tích kiểu báo cáo cũ.
    all_headers = []
    for row in range(1, 5):
        all_headers += [str(c.value) for c in ws[row] if c.value]
    joined = " ".join(all_headers)
    assert "Điểm bất thường" not in joined
    assert "Mức độ" not in joined

    # Mỗi nhà thầu có cột 'ĐG tổng hợp' và 'Thành tiền NT chào' riêng (4 block).
    cols = _leaf_columns(ws)
    assert len(cols.get("ĐG tổng hợp", [])) == 4
    assert len(cols.get("Thành tiền NT chào", [])) == 4


def test_summary_marks_deviating_price_cells_directly(tmp_path: Path):
    prices = {"NT A": 100, "NT B": 105, "NT C": 95, "NT Lệch": 500}
    summary_path = _build_summary(tmp_path, prices)

    wb = load_workbook(summary_path)
    ws = wb["1. HT điện"]
    cols = _leaf_columns(ws)
    dg_cols = sorted(cols["ĐG tổng hợp"])

    data_row = None
    for row in range(5, ws.max_row + 1):
        if any(isinstance(ws.cell(row, c).value, (int, float)) for c in dg_cols):
            data_row = row
            break
    assert data_row is not None

    flagged = []
    for c in dg_cols:
        cell = ws.cell(data_row, c)
        has_fill = cell.fill is not None and cell.fill.fgColor.rgb not in (None, "00000000")
        if cell.comment is not None:
            flagged.append((cell.value, cell.comment.text, has_fill))

    # Đúng một nhà thầu (giá 500) bị đánh dấu, kèm màu nền và ghi chú.
    assert len(flagged) == 1
    value, comment_text, has_fill = flagged[0]
    assert value == 500
    assert has_fill
    assert "cao hơn" in comment_text


def test_summary_no_marks_when_prices_close(tmp_path: Path):
    prices = {"NT A": 100, "NT B": 102, "NT C": 98}
    summary_path = _build_summary(tmp_path, prices)

    wb = load_workbook(summary_path)
    ws = wb["1. HT điện"]
    cols = _leaf_columns(ws)
    dg_cols = sorted(cols["ĐG tổng hợp"])

    for row in range(5, ws.max_row + 1):
        for c in dg_cols:
            cell = ws.cell(row, c)
            if isinstance(cell.value, (int, float)):
                assert cell.comment is None


def test_summary_splits_multiple_hangmuc_into_separate_sheets(tmp_path: Path):
    # Mỗi nhà thầu có 2 sheet hạng mục khác nhau -> file tổng hợp phải có 2 sheet.
    bidder_files = []
    for name, price in {"NT A": 100, "NT B": 110}.items():
        path = tmp_path / f"{name}.xlsx"
        wb = Workbook()
        ws1 = wb.active
        ws1.title = "1. HT điện"
        ws1.append(["STT", "Mã hiệu", "Tên hạng mục", "ĐVT", "Khối lượng", "Đơn giá tổng hợp", "Thành tiền"])
        ws1.append(["1", "M-01", "Tủ điện tổng", "Cái", 1, price, price])
        ws2 = wb.create_sheet("3. HT CTN")
        ws2.append(["STT", "Mã hiệu", "Tên hạng mục", "ĐVT", "Khối lượng", "Đơn giá tổng hợp", "Thành tiền"])
        ws2.append(["1", "P-01", "Máy bơm nước", "Bộ", 1, price * 10, price * 10])
        wb.save(path)
        bidder_files.append((name, path))

    result = compare_bidder_files(bidder_files, config=_cfg())
    summary_path = tmp_path / "Bang_tong_hop.xlsx"
    export_consolidated_summary(result, summary_path)

    wb = load_workbook(summary_path)
    assert "1. HT điện" in wb.sheetnames
    assert "3. HT CTN" in wb.sheetnames

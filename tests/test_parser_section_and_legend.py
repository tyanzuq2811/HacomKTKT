"""Regression: dòng đánh số cột và tiêu đề mục có subtotal không được coi là
hạng mục thật (tránh sinh cờ 'phát sinh'/'bất thường' giả trong báo cáo).

Bắt nguồn từ file chào giá Hacom thực tế:
- Dòng 8 là chú giải đánh số cột 1,2,3,... (Calamine trả về dạng float 1.0, 2.0).
- Dòng "A. ĐẦU MỤC CÔNG VIỆC THEO KLMT" là tổng phụ của mục (chỉ có thành tiền,
  không có đơn vị/khối lượng/đơn giá).
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook

from core.excel_reader import _is_numbering_row, load_workbook_items
from core.models import DocumentRole, RowType


def test_numbering_row_with_float_values_is_detected():
    # Calamine trả số nguyên dưới dạng float: "1.0", "2.0"...
    row = [1.0, 2.0, "3", 4.0, 5.0, 6.0, 7.0, 8.0, "16=11+12+13+14+15"]
    assert _is_numbering_row(row) is True


def test_numbering_row_does_not_misfire_on_real_priced_row():
    # Một dòng hạng mục thật bắt đầu bằng 1,2,3,4 nhưng có chữ -> không phải legend.
    row = [1.0, "Tủ điện tổng", "Cái", 4.0, "Schneider"]
    assert _is_numbering_row(row) is False


def _book_with_legend_and_section(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "1. HT điện"
    ws.append(["STT", "Mã hiệu", "Tên hạng mục", "ĐVT", "Khối lượng", "Đơn giá tổng hợp", "Thành tiền"])
    # Dòng chú giải đánh số cột (số nguyên -> openpyxl/Excel lưu dạng số).
    ws.append([1, 2, 3, 4, 5, 6, 7])
    # Tiêu đề mục cấp cao kèm subtotal, không có đơn vị/khối lượng/đơn giá.
    ws.append(["A", None, "ĐẦU MỤC CÔNG VIỆC THEO KLMT", None, None, None, 76_000_000_000])
    # Hạng mục thật.
    ws.append(["1", "M-01", "Tủ điện tổng", "Cái", 1, 1_000_000, 1_000_000])
    wb.save(path)


def test_legend_and_section_subtotal_excluded_from_comparable_items(tmp_path: Path):
    path = tmp_path / "bidder.xlsx"
    _book_with_legend_and_section(path)

    wb = load_workbook_items(path, DocumentRole.HSDT, bidder="NT A")

    names = {it.item_name for it in wb.items}
    # Dòng đánh số cột không được xuất hiện như một hạng mục (tên "2"/"2.0").
    assert "2" not in names and "2.0" not in names

    # Tiêu đề mục "A. ĐẦU MỤC..." được giữ lại nhưng KHÔNG phải dòng so sánh.
    section = next((it for it in wb.items if "ĐẦU MỤC" in it.item_name), None)
    assert section is not None
    assert section.row_type is RowType.SUMMARY
    assert section.is_comparable is False

    # Chỉ còn đúng một hạng mục thật để so sánh.
    comparable = [it for it in wb.items if it.is_comparable]
    assert len(comparable) == 1
    assert comparable[0].item_name == "Tủ điện tổng"

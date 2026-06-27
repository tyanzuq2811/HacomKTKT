"""Kiểm tra sheet 'Ma trận đơn giá' đánh dấu trực tiếp lên ô giá bị chênh
lệch nhiều giữa các nhà thầu - tô màu ô đó VÀ gắn ghi chú (comment) ngay
trên chính ô đó, không thêm cột riêng.
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook, load_workbook

from core.config import EnterpriseConfig
from core.pipeline import compare_bidder_files
from core.reporter import export_comparison_report


def _cfg() -> EnterpriseConfig:
    cfg = EnterpriseConfig()
    cfg.enable_semantic_matching = False
    cfg.enable_reranker = False
    return cfg


def _bidder_book(path: Path, price: float) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "BOQ"
    ws.append(["STT", "Mã hiệu", "Tên hạng mục", "ĐVT", "Khối lượng", "Đơn giá tổng hợp", "Thành tiền"])
    ws.append(["1", "M-01", "Tủ điện tổng", "Cái", 1, price, price])
    wb.save(path)


def test_price_matrix_marks_the_outlier_cell_directly(tmp_path: Path):
    # Ba nhà thầu báo giá gần nhau (100, 105, 95) và một nhà thầu báo giá
    # cách biệt rất xa (500) cho đúng một hạng mục giống nhau.
    prices = {"NT Gần 1": 100, "NT Gần 2": 105, "NT Gần 3": 95, "NT Lệch xa": 500}
    bidder_files = []
    for name, price in prices.items():
        path = tmp_path / f"{name}.xlsx"
        _bidder_book(path, price)
        bidder_files.append((name, path))

    result = compare_bidder_files(bidder_files, config=_cfg())
    report_path = tmp_path / "report.xlsx"
    export_comparison_report(result, report_path)

    wb = load_workbook(report_path)
    assert "Ma trận đơn giá" in wb.sheetnames
    ws = wb["Ma trận đơn giá"]

    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    # Không có cột riêng cho ghi chú chênh lệch - việc đánh dấu nằm ngay trên ô giá.
    assert "Ghi chú chênh lệch" not in headers

    outlier_col_index = headers.index("NT Lệch xa") + 1
    outlier_cell = ws.cell(row=2, column=outlier_col_index)
    # Ô giá của nhà thầu lệch xa phải được tô màu cảnh báo/nghiêm trọng...
    assert outlier_cell.fill is not None
    assert outlier_cell.fill.fgColor.rgb not in (None, "00000000")
    # ...và có ghi chú (comment) đính trực tiếp trên ô, giải thích rõ mức lệch.
    assert outlier_cell.comment is not None
    assert "NT Lệch xa" in outlier_cell.comment.text
    assert "cao hơn" in outlier_cell.comment.text

    close_bidder_col_index = headers.index("NT Gần 1") + 1
    normal_cell = ws.cell(row=2, column=close_bidder_col_index)
    # Giá gần trung vị thì không bị tô màu cảnh báo và không có ghi chú.
    assert normal_cell.fill is None or normal_cell.fill.fgColor.rgb in (None, "00000000")
    assert normal_cell.comment is None


def test_price_matrix_has_no_comment_when_all_bidders_close(tmp_path: Path):
    prices = {"NT A": 100, "NT B": 102, "NT C": 98}
    bidder_files = []
    for name, price in prices.items():
        path = tmp_path / f"{name}.xlsx"
        _bidder_book(path, price)
        bidder_files.append((name, path))

    result = compare_bidder_files(bidder_files, config=_cfg())
    report_path = tmp_path / "report.xlsx"
    export_comparison_report(result, report_path)

    wb = load_workbook(report_path)
    ws = wb["Ma trận đơn giá"]
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    for bidder in ("NT A", "NT B", "NT C"):
        col_index = headers.index(bidder) + 1
        cell = ws.cell(row=2, column=col_index)
        assert cell.comment is None

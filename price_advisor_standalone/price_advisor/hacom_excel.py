from __future__ import annotations

import csv
import hashlib
import logging
import re
import unicodedata
from dataclasses import dataclass
from datetime import date
from pathlib import Path
from typing import Any, Iterable

from openpyxl import load_workbook

from .normalizer import normalize_description, normalize_unit
from .schemas import PriceReference

log = logging.getLogger(__name__)

SUMMARY_SHEET_NAMES = {"tong hop", "tong", "tong vk", "bia", "dieu khoan", "pham vi tinh toan", "dmvt"}
DEFAULT_QUOTE_DIR = Path("5. Tong hop chao gia 11.12.2025")

# Outlier detection thresholds
PRICE_OUTLIER_LOW = 500          # VND — giá dưới mức này gần chắc là lỗi hoặc phụ kiện rất nhỏ
PRICE_OUTLIER_HIGH = 5_000_000_000  # 5 tỷ VND — giá trên mức này cần kiểm tra

_DATE_PATTERN = re.compile(r"(\d{4})\.(\d{2})\.(\d{2})")


def _parse_date_from_filename(path: Path) -> date | None:
    """Extract date from filenames like '2025.12.08 Chao gia ME ...'."""
    match = _DATE_PATTERN.search(path.name)
    if match:
        try:
            return date(int(match.group(1)), int(match.group(2)), int(match.group(3)))
        except ValueError:
            return None
    return None


def _ascii(text: Any) -> str:
    value = unicodedata.normalize("NFD", str(text or "").replace("\n", " ").lower())
    value = "".join(char for char in value if unicodedata.category(char) != "Mn")
    value = value.replace("đ", "d")
    return re.sub(r"\s+", " ", value).strip()


def _is_number(value: Any) -> bool:
    if isinstance(value, bool) or value is None:
        return False
    if isinstance(value, (int, float)):
        return float(value) > 0
    try:
        return float(str(value).replace(",", "").strip()) > 0
    except ValueError:
        return False


def _to_float(value: Any) -> float | None:
    if not _is_number(value):
        return None
    return float(value)


def _clean_text(value: Any) -> str:
    return str(value or "").replace("\n", " ").strip()


def _safe_ref_id(parts: Iterable[Any]) -> str:
    raw = "|".join(_clean_text(part) for part in parts)
    digest = hashlib.sha1(raw.encode("utf-8")).hexdigest()[:12]
    return f"HACOM-{digest}"


def _bidder_from_filename(path: Path) -> str:
    name = path.stem
    lower = _ascii(name)
    if "linh anh" in lower:
        return "Linh Anh"
    if "van lang" in lower or "tri trung" in lower:
        return "Văn Lang - Trí Trung"
    if "searefico" in lower:
        return "Searefico"
    if "van khanh" in lower:
        return "Vân Khánh"
    return re.sub(r"^\d+\.\s*", "", name).strip()


@dataclass(slots=True)
class HacomSheetLayout:
    header_row: int
    description_col: int
    unit_col: int
    quantity_col: int
    material_col: int
    model_col: int
    brand_col: int
    origin_col: int
    price_main_col: int
    price_aux_col: int
    price_labor_col: int
    price_management_col: int
    price_profit_col: int
    unit_price_col: int
    amount_col: int


def _find_layout(ws) -> HacomSheetLayout | None:
    for row_index, row in enumerate(ws.iter_rows(min_row=1, max_row=min(ws.max_row, 12), values_only=True), start=1):
        normalized = {col_index: _ascii(value) for col_index, value in enumerate(row, start=1)}
        amount_cols = [col for col, text in normalized.items() if "thanh tien boq" in text]
        qty_cols = [col for col, text in normalized.items() if "kl nha thau" in text]
        if not amount_cols or not qty_cols:
            continue

        description_col = 0
        for col, text in normalized.items():
            if any(key in text for key in ("dien giai", "noi dung cong viec", "danh muc quat")):
                description_col = col
                break
        if not description_col:
            continue

        unit_col = 0
        for col, text in normalized.items():
            if "don vi" in text and description_col < col < qty_cols[0]:
                unit_col = col
                break
        if not unit_col:
            unit_col = description_col + 1

        quantity_col = qty_cols[0]
        amount_col = amount_cols[0]
        unit_price_col = amount_col - 1
        price_profit_col = unit_price_col - 1
        price_management_col = unit_price_col - 2
        price_labor_col = unit_price_col - 3
        price_aux_col = unit_price_col - 4
        price_main_col = unit_price_col - 5
        material_col = quantity_col + 1
        return HacomSheetLayout(
            header_row=row_index,
            description_col=description_col,
            unit_col=unit_col,
            quantity_col=quantity_col,
            material_col=material_col,
            model_col=material_col + 1,
            brand_col=material_col + 2,
            origin_col=material_col + 3,
            price_main_col=price_main_col,
            price_aux_col=price_aux_col,
            price_labor_col=price_labor_col,
            price_management_col=price_management_col,
            price_profit_col=price_profit_col,
            unit_price_col=unit_price_col,
            amount_col=amount_col,
        )
    return None


def _is_detail_row(stt: str, description: str, unit: str, price: float | None) -> bool:
    if price is None or price <= 0:
        return False
    if not description or not unit:
        return False
    normalized_description = normalize_description(description)
    if len(normalized_description) < 3:
        return False
    if normalized_description in {
        "dau cong viec theo klmt",
        "hang muc dien",
        "hang muc dien nhe",
        "hang muc cap thoat nuoc",
        "he thong hvac",
    }:
        return False
    # Group rows are usually A/I/II/1 with blank price; this keeps the filter
    # conservative if a formatted total row still has a price.
    if len(stt) <= 3 and description.isupper() and " " in description:
        return False
    return True


def extract_hacom_references_from_workbook(path: str | Path, *, bidder: str | None = None) -> list[PriceReference]:
    workbook_path = Path(path)
    bidder_name = bidder or _bidder_from_filename(workbook_path)
    observed_at = _parse_date_from_filename(workbook_path)
    refs: list[PriceReference] = []
    wb = load_workbook(workbook_path, read_only=True, data_only=True)

    for ws in wb.worksheets:
        if _ascii(ws.title) in SUMMARY_SHEET_NAMES:
            continue
        layout = _find_layout(ws)
        if layout is None:
            continue

        for row_index, row in enumerate(
            ws.iter_rows(min_row=layout.header_row + 2, max_row=ws.max_row, values_only=True),
            start=layout.header_row + 2,
        ):
            values = {index + 1: value for index, value in enumerate(row)}
            stt = _clean_text(values.get(1))
            description = _clean_text(values.get(layout.description_col))
            unit = normalize_unit(_clean_text(values.get(layout.unit_col)))
            price = _to_float(values.get(layout.unit_price_col))
            if not _is_detail_row(stt, description, unit, price):
                continue

            quantity = _to_float(values.get(layout.quantity_col))
            amount = _to_float(values.get(layout.amount_col))
            material = _clean_text(values.get(layout.material_col))
            model = _clean_text(values.get(layout.model_col))
            brand = _clean_text(values.get(layout.brand_col))
            origin = _clean_text(values.get(layout.origin_col))
            enriched_description = description
            if material and normalize_description(material) not in normalize_description(description):
                enriched_description = f"{description} | {material}"

            ref_id = _safe_ref_id([workbook_path.name, ws.title, row_index, bidder_name, stt, description, price])
            refs.append(
                PriceReference(
                    ref_id=ref_id,
                    description=enriched_description,
                    unit=unit,
                    price=price,
                    source=f"HACOM Mall - {bidder_name}",
                    source_type="historical_bid",
                    observed_at=observed_at,
                    metadata={
                        "bidder": bidder_name,
                        "workbook": workbook_path.name,
                        "sheet": ws.title,
                        "row_number": row_index,
                        "stt": stt,
                        "quantity": quantity or 0,
                        "amount": amount or 0,
                        "material": material,
                        "model": model,
                        "brand": brand,
                        "origin": origin,
                        "price_main": _to_float(values.get(layout.price_main_col)) or 0,
                        "price_aux": _to_float(values.get(layout.price_aux_col)) or 0,
                        "price_labor": _to_float(values.get(layout.price_labor_col)) or 0,
                        "price_management": _to_float(values.get(layout.price_management_col)) or 0,
                        "price_profit": _to_float(values.get(layout.price_profit_col)) or 0,
                    },
                )
            )
    return refs


def find_hacom_quote_workbooks(data_dir: str | Path) -> list[Path]:
    root = Path(data_dir)
    quote_dir = root / DEFAULT_QUOTE_DIR
    search_root = quote_dir if quote_dir.exists() else root
    files = []
    for path in sorted(search_root.glob("*.xlsx")):
        lower = _ascii(path.name)
        if path.name.startswith("~$"):
            continue
        if lower.startswith("0.") or "tong hop chao gia" in lower:
            continue
        if "chao gia" in lower or "van khanh" in lower:
            files.append(path)
    return files


def extract_hacom_references(data_dir: str | Path) -> list[PriceReference]:
    refs: list[PriceReference] = []
    for workbook_path in find_hacom_quote_workbooks(data_dir):
        refs.extend(extract_hacom_references_from_workbook(workbook_path))
    _log_price_outliers(refs)
    return refs


def _log_price_outliers(refs: list[PriceReference]) -> None:
    """Log warnings for prices that look suspiciously low or high."""
    low_outliers: list[tuple[str, str, float]] = []
    high_outliers: list[tuple[str, str, float]] = []
    for ref in refs:
        if ref.price <= PRICE_OUTLIER_LOW:
            low_outliers.append((ref.ref_id, ref.description[:60], ref.price))
        elif ref.price >= PRICE_OUTLIER_HIGH:
            high_outliers.append((ref.ref_id, ref.description[:60], ref.price))
    if low_outliers:
        log.warning(
            "Phát hiện %d dòng có giá rất thấp (≤ %s VND) — có thể là phụ kiện nhỏ hoặc lỗi dữ liệu:",
            len(low_outliers), f"{PRICE_OUTLIER_LOW:,}",
        )
        for ref_id, desc, price in low_outliers[:10]:
            log.warning("  %s | %s | %s VND", ref_id, desc, f"{price:,.0f}")
        if len(low_outliers) > 10:
            log.warning("  ... và %d dòng nữa", len(low_outliers) - 10)
    if high_outliers:
        log.warning(
            "Phát hiện %d dòng có giá rất cao (≥ %s VND) — có thể là tủ điện tổng hoặc thiết bị lớn:",
            len(high_outliers), f"{PRICE_OUTLIER_HIGH:,}",
        )
        for ref_id, desc, price in high_outliers[:10]:
            log.warning("  %s | %s | %s VND", ref_id, desc, f"{price:,.0f}")
        if len(high_outliers) > 10:
            log.warning("  ... và %d dòng nữa", len(high_outliers) - 10)


def export_references_csv(refs: Iterable[PriceReference], output: str | Path) -> int:
    output_path = Path(output)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    fieldnames = [
        "ref_id",
        "description",
        "unit",
        "price",
        "source",
        "source_type",
        "observed_at",
        "bidder",
        "workbook",
        "sheet",
        "row_number",
        "stt",
        "quantity",
        "amount",
        "material",
        "model",
        "brand",
        "origin",
        "price_main",
        "price_aux",
        "price_labor",
        "price_management",
        "price_profit",
    ]
    rows = list(refs)
    with output_path.open("w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        for ref in rows:
            row = {
                "ref_id": ref.ref_id,
                "description": ref.description,
                "unit": ref.unit,
                "price": ref.price,
                "source": ref.source,
                "source_type": ref.source_type,
                "observed_at": ref.observed_at.isoformat() if ref.observed_at else "",
            }
            row.update({key: ref.metadata.get(key, "") for key in fieldnames if key not in row})
            writer.writerow(row)
    return len(rows)


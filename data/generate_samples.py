"""
Tạo file Excel mẫu thực tế để test pipeline.
Chạy: python data/generate_samples.py
"""
import sys
from pathlib import Path
sys.path.insert(0, str(Path(__file__).parent.parent))

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

def _hdr_style(ws, row, cols, values, bg="1A3A5C"):
    for col, val in zip(cols, values):
        c = ws.cell(row=row, column=col, value=val)
        c.fill = PatternFill("solid", fgColor=bg)
        c.font = Font(color="FFFFFF", bold=True, size=10, name="Calibri")
        c.alignment = Alignment(horizontal="center", vertical="center")
        bd = Side(style="thin", color="999999")
        c.border = Border(left=bd, right=bd, top=bd, bottom=bd)

def _data_cell(ws, row, col, val):
    c = ws.cell(row=row, column=col, value=val)
    bd = Side(style="thin", color="D1D5DB")
    c.border = Border(left=bd, right=bd, top=bd, bottom=bd)
    c.alignment = Alignment(vertical="center")
    return c

# ── Dữ liệu HSMT ─────────────────────────────────────────────────
HSMT_DATA = [
    # ma_hieu, ten_hang_muc, dvt, kl, don_gia, thanh_tien, vat_tu_spec
    ("M-01", "Ống thép mạ kẽm DN100 PN16", "m", 120, 850_000, 102_000_000, "ASTM A53 Grade B, mạ kẽm nhúng nóng"),
    ("M-02", "Ống thép mạ kẽm DN80 PN16", "m", 85, 620_000, 52_700_000, "ASTM A53 Grade B, mạ kẽm nhúng nóng"),
    ("M-03", "Co 90° DN100 mạ kẽm", "cái", 24, 180_000, 4_320_000, "Đúc gang, mạ kẽm, PN16"),
    ("M-04", "Ống thép DN150 mạ kẽm PN16", "m", 60, 1_100_000, 66_000_000, "ASTM A53 Grade B, mạ kẽm nhúng nóng"),
    ("M-05", "Bích mù DN150 PN16", "cái", 8, 250_000, 2_000_000, "Gang đúc, PN16 ANSI B16.5"),
    ("M-06", "Van cổng DN100 PN16", "cái", 12, 1_800_000, 21_600_000, "Gang cầu, PN16, tay quay"),
    ("M-07", "Van 1 chiều DN80", "cái", 6, 950_000, 5_700_000, "Swing check, gang đúc PN10"),
    ("M-08", "Co 45° DN80 mạ kẽm", "cái", 18, 120_000, 2_160_000, "Đúc gang, mạ kẽm, PN16"),
    ("M-09", "Van bướm DN200 PN16", "cái", 6, 3_200_000, 19_200_000, "Gang cầu, tay đòn, PN16"),
    ("M-10", "Ống thép đen DN200", "m", 45, 1_950_000, 87_750_000, "ASTM A53 Grade B, sơn epoxy"),
    ("C-01", "Bê tông M200 đài cọc", "m³", 120, 1_800_000, 216_000_000, "Bê tông thương phẩm, đổ tại chỗ"),
    ("C-02", "Bê tông M300 dầm sàn", "m³", 84, 2_100_000, 176_400_000, "Bê tông thương phẩm, đổ tại chỗ"),
    ("C-03", "Cốt thép CB300-V", "tấn", 18.5, 18_500_000, 342_250_000, "Thép cán nóng CB300-V, TCVN 1651-2:2018"),
    ("C-04", "Ván khuôn thép tấm", "m²", 450, 95_000, 42_750_000, "Ván khuôn định hình, t=3mm"),
    ("E-01", "Tủ điện chính MSB-01", "bộ", 1, 85_000_000, 85_000_000, "Tủ điện ngoài trời IP54, CB tổng 400A"),
    ("E-02", "Tủ điện phân phối DB-01", "bộ", 3, 22_000_000, 66_000_000, "Tủ điện trong nhà IP31, CB 100A"),
    ("E-03", "Cáp điện lực 3×16mm² + 10mm²", "m", 320, 95_000, 30_400_000, "CXV/DSTA, 0.6/1kV, Cadivi"),
    ("E-04", "Cáp điện 2×2.5mm²", "m", 850, 18_000, 15_300_000, "CVVs, 450/750V, Cadivi"),
    ("E-05", "Đèn LED panel 600×600 40W", "bộ", 48, 420_000, 20_160_000, "LED 40W, 4000K, IP44, Philips/Rạng Đông"),
    ("E-06", "Ổ cắm 3 chấu 16A", "cái", 120, 45_000, 5_400_000, "IP44, chịu nước, Schneider/Legrand"),
    ("E-07", "Ống luồn dây PVC Ø25", "m", 600, 12_000, 7_200_000, "Ống nhựa PVC cứng, Tiền Phong"),
    ("E-11", "Tủ điện MCC tổng", "bộ", 1, 206_000_000, 206_000_000, "Motor Control Center, IP54, Siemens/ABB"),
]

# ── Dữ liệu HSDT nhà thầu (có sai lệch cố tình) ─────────────────
HSDT_MINHPHAT = [
    # ma_hieu, ten_hang_muc, dvt, kl, don_gia, thanh_tien, vat_tu, ghi_chu
    ("M-01", "Ống thép mạ kẽm DN100 PN16", "m", 120, 890_000, 106_800_000, "ASTM A53 Grade B, mạ kẽm nhúng nóng", ""),
    ("M-02", "Ống thép mạ kẽm DN80 PN16", "m", 85, 618_000, 52_530_000, "ASTM A53 Grade B", ""),
    ("M-03", "Co 90° DN100 mạ kẽm", "cái", 24, 175_000, 4_200_000, "Gang đúc mạ kẽm PN16", ""),
    ("M-04", "Ống thép đen DN150",  "m", 60, 980_000, 58_800_000, "Carbon steel pipe DN150",
     "⚠ SAI SPEC: Chào ống đen, HSMT yêu cầu mạ kẽm"),
    ("M-05", "Bích mù DN150 PN16", "cái", 8, 252_000, 2_016_000, "Gang đúc PN16", ""),
    ("M-06", "Van cổng DN100 PN16", "cái", 12, 1_820_000, 21_840_000, "Gang cầu PN16", ""),
    ("M-07", "Van 1 chiều DN80", "cái", 6, 945_000, 5_670_000, "Swing check PN10", ""),
    ("M-08", "Co 45° DN80 mạ kẽm", "cái", 18, 118_000, 2_124_000, "Gang đúc mạ kẽm", ""),
    ("M-09", "VB-DN200-16BAR",      "cái", 6, 3_150_000, 18_900_000, "Gang cầu PN16",
     "Mã hiệu khác: VB200-16BAR"),
    ("M-10", "Ống thép đen DN200", "m", 45, 1_920_000, 86_400_000, "ASTM A53 sơn epoxy", ""),
    ("C-01", "Bê tông M200 đài cọc", "m³", 120, 1_780_000, 213_600_000, "Bê tông thương phẩm", ""),
    ("C-02", "Bê tông M300 dầm sàn", "m³", 72, 2_080_000, 149_760_000, "Bê tông thương phẩm",
     "⚠ Khối lượng tính thiếu 12m³"),
    ("C-03", "Cốt thép CB300-V", "tấn", 18.5, 18_200_000, 336_700_000, "Thép CB300-V TCVN", ""),
    ("C-04", "Ván khuôn thép tấm", "m²", 450, 93_000, 41_850_000, "Ván khuôn định hình t=3mm", ""),
    ("E-01", "Tủ điện chính MSB-01", "bộ", 1, 84_000_000, 84_000_000, "Tủ điện ngoài trời IP54", ""),
    ("E-02", "Tủ điện phân phối DB-01", "bộ", 3, 21_500_000, 64_500_000, "IP31 CB100A", ""),
    ("E-03", "Cáp điện 3×16mm²+10mm²", "m", 320, 94_000, 30_080_000, "CXV/DSTA 0.6/1kV Cadivi", ""),
    ("E-04", "Cáp điện 2×2.5mm²", "m", 850, 17_800_000, 15_130_000, "CVVs Cadivi",
     "⚠ LỖI TOÁN: Thành tiền sai (850×17800=15,130,000 đúng; nhà thầu ghi nhầm đơn vị)"),
    ("E-05", "Đèn LED panel 600×600 40W", "bộ", 50, 415_000, 20_750_000, "LED 40W 4000K Rạng Đông", ""),
    ("E-06", "Ổ cắm 3 chấu 16A IP44", "cái", 120, 44_000, 5_280_000, "IP44 Schneider", ""),
    ("E-07", "Ống luồn dây PVC Ø25", "m", 600, 11_800, 7_080_000, "PVC cứng Tiền Phong", ""),
    ("E-11", "Tủ điện MCC tổng", "bộ", 1, 285_000_000, 285_000_000, "MCC IP54 Siemens",
     "Giá cao hơn HSMT 38% — cần bảng phân tích chi phí"),
]


def make_hsmt(path: str):
    wb = Workbook()
    ws = wb.active
    ws.title = "BOQ_HSMT"

    # Dòng tiêu đề lớn
    ws.merge_cells("A1:G1")
    c = ws["A1"]
    c.value = "HỒ SƠ MỜI THẦU — BẢNG KHỐI LƯỢNG MỜI THẦU — Gói M01"
    c.font = Font(bold=True, size=13, color="FFFFFF", name="Calibri")
    c.fill = PatternFill("solid", fgColor="1A3A5C")
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    headers = ["Mã hiệu", "Tên hạng mục", "ĐVT", "Khối lượng",
               "Đơn giá (đ)", "Thành tiền (đ)", "Vật tư / Spec"]
    _hdr_style(ws, 2, range(1, 8), headers)
    ws.row_dimensions[2].height = 22

    col_widths = [12, 42, 8, 13, 18, 18, 45]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[chr(64+i)].width = w

    for r, row in enumerate(HSMT_DATA, start=3):
        for c_idx, val in enumerate(row, start=1):
            _data_cell(ws, r, c_idx, val)
        ws.row_dimensions[r].height = 16

    wb.save(path)
    print(f"✓ HSMT mẫu: {path}")


def make_hsdt_minhphat(path: str):
    wb = Workbook()
    ws = wb.active
    ws.title = "HSDT_NT_MinhPhat"

    ws.merge_cells("A1:H1")
    c = ws["A1"]
    c.value = "HỒ SƠ DỰ THẦU — CÔNG TY XÂY DỰNG MINH PHÁT — Gói M01"
    c.font = Font(bold=True, size=13, color="FFFFFF", name="Calibri")
    c.fill = PatternFill("solid", fgColor="0F4C2A")
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    headers = ["Mã hiệu", "Tên hạng mục", "ĐVT", "Khối lượng",
               "Đơn giá chào (đ)", "Thành tiền (đ)", "Vật tư / Spec", "Ghi chú"]
    _hdr_style(ws, 2, range(1, 9), headers, bg="0F4C2A")
    ws.row_dimensions[2].height = 22

    col_widths = [14, 42, 8, 13, 18, 18, 45, 42]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[chr(64+i)].width = w

    for r, row in enumerate(HSDT_MINHPHAT, start=3):
        for c_idx, val in enumerate(row, start=1):
            _data_cell(ws, r, c_idx, val)
        ws.row_dimensions[r].height = 16

    wb.save(path)
    print(f"✓ HSDT mẫu (NT Minh Phát): {path}")


if __name__ == "__main__":
    out_dir = Path(__file__).parent
    make_hsmt(str(out_dir / "HSMT_GoiM01.xlsx"))
    make_hsdt_minhphat(str(out_dir / "NT_MinhPhat_HSDT.xlsx"))
    print("\nHoàn thành — kiểm tra thư mục data/")
